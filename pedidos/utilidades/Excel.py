# -*- encoding: utf-8 -*-

from datetime import datetime,date
from catalogo.models import *
from openpyxl import Workbook, workbook

from openpyxl.styles import Font, NamedStyle

from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import os


now = datetime.now()


class Excel(object):
    
    """
        Funcion de utilidad para generar un excel formato generico de pedido
        Parametros:
            productos: lista de productos a crear en el fichero 
            cliente: objeto CustomerBD
            pedido: objeto BD Orders
        
    """   
    def generaExcelPedido(self, productos, cliente, pedido):
        
        self.encabezadoValores = ['Numero pedido','Cliente','Producto', 'Cantidad','Importe','Total']
        self.celdas =['A', 'B', 'C', 'D', 'E', 'F', 'G']
        
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        greyFill = PatternFill(start_color='A4A4A4',end_color='A4A4A4',fill_type='solid') 
        
        #DIMENSIONES DE COLUMNAS
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['c'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
         
        fila = 1
        
        posicion = 0
        longitud = len( self.encabezadoValores)
                
        for celda in self.celdas[0:longitud]:
            if posicion < longitud:
                ws[celda + str(fila)] = self.encabezadoValores[posicion]
                ws[celda + str(fila)].font = Font(bold = True) 
                ws[celda + str(fila)].fill = greyFill
                
            posicion = posicion + 1
        fila = fila + 1
            
        
        for prod in productos:
            ws.cell(row=fila,column=1).value = pedido.number
            ws.cell(row=fila,column=2).value = cliente.name + " " + cliente.lastname
            ws.cell(row=fila,column=3).value = prod['name']
            ws.cell(row=fila,column=4).value = prod['cantidad']
            ws.cell(row=fila,column=5).value = prod['price']
            ws.cell(row=fila,column=5).value = prod['total']
                    
            fila = fila + 1
        
        #Establecemos el nombre del archivo
        nombrearchivo = "pedido.xlsx"
                
        directory = 'media/pedidos/' + str(cliente.pk) + '/' + pedido.number + "/"
                
        if not os.path.exists(directory):
            os.makedirs(directory)
            ruta = directory + '/' + nombrearchivo
            wb.save(ruta)           
        else:
            ruta = directory + '/' + nombrearchivo
            wb.save(ruta)
            
        return nombrearchivo